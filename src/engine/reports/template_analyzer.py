"""
Template Analyzer - Auto-detect structure from Excel templates.

Detects:
- Tables, ranges, and data regions
- Header rows and column types (dimension vs KPI)
- Pivot structures and aggregation patterns
- Placeholder patterns for dynamic population
"""

import re
import logging
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

logger = logging.getLogger(__name__)


@dataclass
class DetectedTable:
    """Represents a detected table in a worksheet."""
    sheet_name: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    headers: List[str]
    header_row: int
    dimension_cols: List[str] = field(default_factory=list)
    kpi_cols: List[str] = field(default_factory=list)
    granularity: str = "unknown"
    is_pivot: bool = False
    row_dimension: Optional[str] = None
    col_dimension: Optional[str] = None
    aggregation_type: str = "sum"


class TemplateAnalyzer:
    """
    Intelligent Excel template analyzer.
    
    Detects table structures, header types, pivot layouts, 
    and aggregation patterns without manual configuration.
    """
    
    # Common dimension column patterns
    DIMENSION_PATTERNS = [
        r'^(date|day|week|month|year|period|time)$',
        r'^(platform|channel|source|network|medium)$',
        r'^(campaign|ad_group|adset|ad|creative)$',
        r'^(device|age|gender|geo|region|country|city)$',
        r'^(funnel|stage|objective|placement)$',
    ]
    
    # Common KPI column patterns
    KPI_PATTERNS = [
        r'^(spend|cost|budget|amount|investment|expenditure)$',
        r'^(impressions?|imps?|views?|reach)$',
        r'^(clicks?|sessions?|visits?)$',
        r'^(conversions?|convs?|leads?|signups?|purchases?)$',
        r'^(revenue|sales|value|income)$',
        r'^(ctr|cpc|cpm|cpa|roas|roi|cvr)$',
        r'^(frequency|freq)$',
    ]
    
    # Placeholder patterns
    PLACEHOLDER_PATTERNS = [
        r'\{\{(\w+)\}\}',  # {{metric}}
        r'\{(\w+)\}',      # {dimension}
        r'\<(\w+)\>',      # <date>
        r'\[(\w+)\]',      # [value]
    ]
    
    def __init__(self):
        self._dimension_regex = [re.compile(p, re.IGNORECASE) for p in self.DIMENSION_PATTERNS]
        self._kpi_regex = [re.compile(p, re.IGNORECASE) for p in self.KPI_PATTERNS]
        self._placeholder_regex = [re.compile(p) for p in self.PLACEHOLDER_PATTERNS]
    
    def analyze(self, template_path: str) -> Dict[str, Any]:
        """
        Analyze an Excel template and return its structure.
        
        Args:
            template_path: Path to Excel template file
            
        Returns:
            Dictionary with detected sheets, tables, and structure
        """
        path = Path(template_path)
        if not path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        wb = load_workbook(template_path, data_only=False)
        
        result = {
            "file": str(path.name),
            "sheets": [],
            "total_tables": 0,
            "placeholders": [],
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_analysis = self._analyze_sheet(ws, sheet_name)
            result["sheets"].append(sheet_analysis)
            result["total_tables"] += len(sheet_analysis["tables"])
            result["placeholders"].extend(sheet_analysis.get("placeholders", []))
        
        wb.close()
        
        logger.info(f"Analyzed template: {path.name}, found {result['total_tables']} tables")
        return result
    
    def _analyze_sheet(self, ws: Worksheet, sheet_name: str) -> Dict[str, Any]:
        """Analyze a single worksheet."""
        tables = []
        placeholders = []
        
        # Find all data regions
        regions = self._find_data_regions(ws)
        
        for region in regions:
            table = self._analyze_region(ws, sheet_name, region)
            if table and len(table.headers) > 0:
                tables.append(self._table_to_dict(table))
        
        # Find placeholders
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for pattern in self._placeholder_regex:
                        matches = pattern.findall(cell.value)
                        for match in matches:
                            placeholders.append({
                                "name": match,
                                "cell": cell.coordinate,
                                "sheet": sheet_name
                            })
        
        return {
            "name": sheet_name,
            "tables": tables,
            "placeholders": placeholders,
        }
    
    def _find_data_regions(self, ws: Worksheet) -> List[Tuple[int, int, int, int]]:
        """
        Find contiguous data regions in a worksheet.
        
        Returns list of (start_row, end_row, start_col, end_col) tuples.
        """
        regions = []
        visited = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if (row_idx, col_idx) in visited:
                    continue
                    
                if cell.value is not None:
                    # Found a cell with data - explore the region
                    region = self._flood_fill_region(ws, row_idx, col_idx, visited)
                    if region:
                        regions.append(region)
        
        return regions
    
    def _flood_fill_region(self, ws: Worksheet, start_row: int, start_col: int, 
                          visited: set) -> Optional[Tuple[int, int, int, int]]:
        """Find the extent of a contiguous data region."""
        min_row, max_row = start_row, start_row
        min_col, max_col = start_col, start_col
        
        # Simple approach: expand from starting point until empty rows/cols
        # Find header row (first row with multiple values)
        header_row = start_row
        for r in range(start_row, min(start_row + 5, ws.max_row + 1)):
            row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            non_empty = sum(1 for v in row_values if v is not None)
            if non_empty >= 2:
                header_row = r
                break
        
        # Find column extent from header row
        for c in range(1, ws.max_column + 1):
            if ws.cell(header_row, c).value is not None:
                min_col = min(min_col, c)
                max_col = max(max_col, c)
        
        # Find row extent
        empty_rows = 0
        for r in range(header_row, ws.max_row + 1):
            row_values = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
            if all(v is None for v in row_values):
                empty_rows += 1
                if empty_rows >= 2:
                    break
            else:
                empty_rows = 0
                max_row = r
        
        # Mark visited
        for r in range(header_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                visited.add((r, c))
        
        # Only return if region has minimum size
        if max_row - header_row >= 1 and max_col - min_col >= 1:
            return (header_row, max_row, min_col, max_col)
        
        return None
    
    def _analyze_region(self, ws: Worksheet, sheet_name: str, 
                       region: Tuple[int, int, int, int]) -> Optional[DetectedTable]:
        """Analyze a data region to extract table structure."""
        start_row, end_row, start_col, end_col = region
        
        # Extract headers from first row
        headers = []
        for col in range(start_col, end_col + 1):
            cell_value = ws.cell(start_row, col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column_{col}")
        
        if not headers:
            return None
        
        # Classify columns
        dimension_cols = []
        kpi_cols = []
        
        for header in headers:
            header_clean = re.sub(r'[^a-zA-Z0-9_]', '', header.lower())
            
            is_dim = any(p.match(header_clean) for p in self._dimension_regex)
            is_kpi = any(p.match(header_clean) for p in self._kpi_regex)
            
            if is_dim:
                dimension_cols.append(header)
            elif is_kpi:
                kpi_cols.append(header)
            else:
                # Heuristic: if it looks numeric in data rows, it's a KPI
                sample_values = [ws.cell(r, start_col + headers.index(header)).value 
                               for r in range(start_row + 1, min(start_row + 5, end_row + 1))]
                if any(isinstance(v, (int, float)) for v in sample_values):
                    kpi_cols.append(header)
                else:
                    dimension_cols.append(header)
        
        # Detect granularity
        granularity = self._detect_granularity(dimension_cols)
        
        # Detect if it's a pivot structure
        is_pivot, row_dim, col_dim = self._detect_pivot(ws, region, headers)
        
        return DetectedTable(
            sheet_name=sheet_name,
            start_row=start_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            headers=headers,
            header_row=start_row,
            dimension_cols=dimension_cols,
            kpi_cols=kpi_cols,
            granularity=granularity,
            is_pivot=is_pivot,
            row_dimension=row_dim,
            col_dimension=col_dim,
        )
    
    def _detect_granularity(self, dimension_cols: List[str]) -> str:
        """Detect the granularity level from dimension columns."""
        dim_lower = [d.lower() for d in dimension_cols]
        
        if any('date' in d or 'day' in d for d in dim_lower):
            return "daily"
        elif any('week' in d for d in dim_lower):
            return "weekly"
        elif any('month' in d for d in dim_lower):
            return "monthly"
        elif any('platform' in d or 'channel' in d for d in dim_lower):
            return "platform"
        elif any('campaign' in d for d in dim_lower):
            return "campaign"
        elif any('device' in d for d in dim_lower):
            return "device"
        elif any('age' in d or 'gender' in d for d in dim_lower):
            return "demographic"
        
        return "aggregate"
    
    def _detect_pivot(self, ws: Worksheet, region: Tuple[int, int, int, int],
                     headers: List[str]) -> Tuple[bool, Optional[str], Optional[str]]:
        """
        Detect if the table is a pivot structure.
        
        A pivot has:
        - Row labels (dimension values) in first column(s)
        - Column labels (another dimension) in header row
        - Values at intersections
        """
        start_row, end_row, start_col, end_col = region
        
        # Check if first column contains repeated categorical values
        first_col_values = [ws.cell(r, start_col).value 
                          for r in range(start_row + 1, end_row + 1)]
        unique_first = set(v for v in first_col_values if v)
        
        # Check if headers (after first) look like dimension values
        other_headers = headers[1:] if len(headers) > 1 else []
        
        # Heuristic: if headers are short strings (like "Meta", "Google", "2024-01")
        # and first column has categorical data, it's likely a pivot
        if len(unique_first) > 1 and len(other_headers) > 1:
            header_lengths = [len(str(h)) for h in other_headers if h]
            if header_lengths and max(header_lengths) < 30:
                return True, headers[0] if headers else None, "columns"
        
        return False, None, None
    
    def _table_to_dict(self, table: DetectedTable) -> Dict[str, Any]:
        """Convert DetectedTable to dictionary."""
        return {
            "start_row": table.start_row,
            "end_row": table.end_row,
            "start_col": table.start_col,
            "end_col": table.end_col,
            "headers": table.headers,
            "header_row": table.header_row,
            "dimension_cols": table.dimension_cols,
            "kpi_cols": table.kpi_cols,
            "granularity": table.granularity,
            "is_pivot": table.is_pivot,
            "row_dimension": table.row_dimension,
            "col_dimension": table.col_dimension,
        }
