"""
Universal Populator - Fill Excel templates with data while preserving formatting.

Features:
- Write data to exact cell positions
- Insert dynamic rows for variable-length data
- Preserve template styles, formatting, and formulas
- Handle both horizontal and vertical layouts
"""

import logging
import os
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, PatternFill
from openpyxl.formula.translate import Translator

logger = logging.getLogger(__name__)


class UniversalPopulator:
    """
    Universal Excel template populator.
    
    Fills any Excel template with data while:
    - Preserving original formatting
    - Handling dynamic row insertion
    - Supporting multiple tables per sheet
    - Maintaining formulas outside data regions
    """
    
    def __init__(self, output_dir: str = "data/pacing_reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def populate(self, template_path: str, 
                 aggregated_data: Dict[str, pd.DataFrame],
                 template_structure: Dict[str, Any],
                 column_mappings: Dict[str, Any] = None,
                 output_name: str = None) -> str:
        """
        Populate template with aggregated data.
        
        Args:
            template_path: Path to Excel template
            aggregated_data: Dict of DataFrames from DynamicAggregator
            template_structure: Output from TemplateAnalyzer
            column_mappings: Output from SmartMapper (optional)
            output_name: Custom output filename (optional)
            
        Returns:
            Path to generated report file
        """
        # Load template
        wb = load_workbook(template_path)
        
        # Process each sheet
        for sheet_info in template_structure.get("sheets", []):
            sheet_name = sheet_info["name"]
            
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet {sheet_name} not found in template")
                continue
            
            ws = wb[sheet_name]
            
            # Process each table in the sheet
            for i, table in enumerate(sheet_info.get("tables", [])):
                table_key = f"{sheet_name}_table_{i}"
                
                if table_key in aggregated_data:
                    df = aggregated_data[table_key]
                    self._populate_table(ws, table, df, column_mappings)
        
        # Fill placeholders
        self._fill_placeholders(wb, template_structure.get("placeholders", []))
        
        # Generate output filename
        if output_name is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"report_{timestamp}.xlsx"
        
        output_path = self.output_dir / output_name
        wb.save(output_path)
        wb.close()
        
        logger.info(f"Generated report: {output_path}")
        return str(output_path)
    
    def _populate_table(self, ws: Worksheet, table: Dict, 
                       df: pd.DataFrame, column_mappings: Dict = None):
        """Populate a single table with data."""
        start_row = table["start_row"]
        start_col = table["start_col"]
        headers = table["headers"]
        
        # Map DataFrame columns to template headers
        col_order = []
        for header in headers:
            header_lower = header.lower().replace(' ', '_')
            
            # Try direct match
            matched_col = None
            for df_col in df.columns:
                if df_col.lower() == header_lower:
                    matched_col = df_col
                    break
            
            # Try via mappings
            if not matched_col and column_mappings:
                mapping = column_mappings.get(header)
                if mapping:
                    data_col = mapping.data_column if hasattr(mapping, 'data_column') else str(mapping)
                    if data_col in df.columns:
                        matched_col = data_col
            
            col_order.append(matched_col)
        
        # Determine if we need to insert rows
        data_rows = len(df)
        template_rows = table["end_row"] - table["start_row"]
        
        if data_rows > template_rows:
            # Insert additional rows
            rows_to_insert = data_rows - template_rows
            self._insert_rows_with_style(ws, table["end_row"], rows_to_insert, start_col, len(headers))
        
        # Write data
        data_start_row = start_row + 1  # Skip header row
        
        for row_idx, (_, row_data) in enumerate(df.iterrows()):
            cell_row = data_start_row + row_idx
            
            for col_idx, df_col in enumerate(col_order):
                cell_col = start_col + col_idx
                
                if df_col and df_col in row_data:
                    value = row_data[df_col]
                    
                    # Handle special types
                    if pd.isna(value):
                        value = 0
                    elif isinstance(value, float):
                        value = round(value, 2)
                    
                    ws.cell(row=cell_row, column=cell_col).value = value
        
        logger.debug(f"Populated table at row {start_row}: {data_rows} rows")
    
    def _insert_rows_with_style(self, ws: Worksheet, insert_after: int, 
                                num_rows: int, start_col: int, num_cols: int):
        """Insert rows while copying style and formulas from template row."""
        # Get style from template row (row above insert point)
        template_row = insert_after
        
        # Insert blank rows
        ws.insert_rows(insert_after + 1, num_rows)
        
        # Copy styles and formulas from template row
        for new_row in range(insert_after + 1, insert_after + 1 + num_rows):
            for col in range(start_col, start_col + num_cols):
                source_cell = ws.cell(row=template_row, column=col)
                target_cell = ws.cell(row=new_row, column=col)
                
                # Copy style
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                
                # Copy and Translate Formula (crucial for Pacing Reports)
                if source_cell.data_type == 'f':
                    formula = source_cell.value
                    try:
                        # Translate relative references (e.g., A10 -> A11)
                        translated_formula = Translator(formula, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
                        target_cell.value = translated_formula
                    except Exception as e:
                        logger.warning(f"Failed to translate formula {formula}: {e}")
                        target_cell.value = formula  # Fallback to original
    
    def _fill_placeholders(self, wb, placeholders: List[Dict]):
        """Fill placeholder values in the workbook."""
        # Default placeholder values
        placeholder_values = {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "datetime": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "report_date": datetime.now().strftime("%Y-%m-%d"),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "today": datetime.now().strftime("%Y-%m-%d"),
        }
        
        for ph in placeholders:
            sheet_name = ph.get("sheet")
            cell_addr = ph.get("cell")
            name = ph.get("name", "").lower()
            
            if sheet_name in wb.sheetnames and cell_addr:
                ws = wb[sheet_name]
                cell = ws[cell_addr]
                
                if name in placeholder_values:
                    # Replace placeholder with value
                    original = str(cell.value) if cell.value else ""
                    for pattern_type in ["{{%s}}", "{%s}", "<%s>", "[%s]"]:
                        pattern = pattern_type % name
                        if pattern in original.lower():
                            original = original.replace(pattern, placeholder_values[name])
                            original = original.replace(pattern.upper(), placeholder_values[name])
                    cell.value = original
    
    def populate_simple(self, template_path: str, data: pd.DataFrame,
                       sheet_name: str = None, start_row: int = 2,
                       output_name: str = None) -> str:
        """
        Simplified population for single-table templates.
        
        Writes DataFrame directly to sheet, matching column names to headers.
        
        Args:
            template_path: Path to template
            data: DataFrame to write
            sheet_name: Target sheet (first sheet if None)
            start_row: Row to start writing data (default 2, assuming 1 is header)
            output_name: Output filename
            
        Returns:
            Path to generated file
        """
        wb = load_workbook(template_path)
        ws = wb.active if sheet_name is None else wb[sheet_name]
        
        # Get headers from row 1
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h]
        
        # Match and write
        for row_idx, (_, row_data) in enumerate(data.iterrows()):
            for col_idx, header in enumerate(headers):
                cell = ws.cell(row=start_row + row_idx, column=col_idx + 1)
                
                # Find matching column in DataFrame
                header_clean = str(header).lower().replace(' ', '_')
                matched = None
                
                for df_col in data.columns:
                    if df_col.lower() == header_clean:
                        matched = df_col
                        break
                
                if matched:
                    value = row_data[matched]
                    if pd.notna(value):
                        cell.value = round(value, 2) if isinstance(value, float) else value
        
        # Save
        if output_name is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"report_{timestamp}.xlsx"
        
        output_path = self.output_dir / output_name
        wb.save(output_path)
        wb.close()
        
        return str(output_path)
