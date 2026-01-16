import pandas as pd
from openpyxl import Workbook
from pathlib import Path
from src.engine.reports.intelligent_engine import IntelligentReportEngine
import shutil
import os

def setup_mock_template(template_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot Analysis"
    
    # Header
    ws.cell(row=1, column=1, value="Channel")
    ws.cell(row=1, column=2, value="Spend")
    ws.cell(row=1, column=3, value="Goal")
    ws.cell(row=1, column=4, value="Variance")
    
    # Template Row (Row 2)
    ws.cell(row=2, column=1, value="Google")
    ws.cell(row=2, column=2, value=100)
    ws.cell(row=2, column=3, value=150)
    # Formula: Variance = Goal - Spend
    ws.cell(row=2, column=4, value="=C2-B2")
    
    wb.save(template_path)
    return template_path

def test_dynamic_pivot_generation():
    template_path = "tests/data/mock_pivot_template.xlsx"
    os.makedirs("tests/data", exist_ok=True)
    setup_mock_template(template_path)
    
    # 3 items in data (template only has 1)
    data = pd.DataFrame([
        {"Channel": "Google", "Spend": 120, "Goal": 150},
        {"Channel": "Meta", "Spend": 80, "Goal": 100},
        {"Channel": "TikTok", "Spend": 50, "Goal": 60}
    ])
    
    output_dir = "tests/outputs"
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    
    engine = IntelligentReportEngine(output_dir=output_dir)
    result = engine.generate(template_path, data, output_name="test_report.xlsx")
    
    output_path = result["output_path"]
    print(f"Report generated at: {output_path}")
    
    # Verification
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb.active
    
    print(f"Max row: {ws.max_row}")
    assert ws.max_row == 4 # Header + 3 data rows
    
    # Check Row 2 (Google)
    assert ws.cell(row=2, column=1).value == "Google"
    assert ws.cell(row=2, column=2).value == 120
    assert ws.cell(row=2, column=4).value == "=C2-B2"
    
    # Check Row 3 (Meta) - Formula should be translated
    assert ws.cell(row=3, column=1).value == "Meta"
    assert ws.cell(row=3, column=4).value == "=C3-B3"
    
    # Check Row 4 (TikTok) - Formula should be translated
    assert ws.cell(row=4, column=1).value == "TikTok"
    assert ws.cell(row=4, column=4).value == "=C4-B4"
    
    print("✓ Dynamic Pivot Analysis verification PASSED!")

if __name__ == "__main__":
    test_dynamic_pivot_generation()
